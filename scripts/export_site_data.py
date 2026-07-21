#!/usr/bin/env python3
"""Export publications.xlsx + OpenAlex citation counts to JSON for the site.

Reads the curated data/publications.xlsx (title, year, venue, type, doi —
respects any manual edits made there) and pulls per-year citation counts from
OpenAlex for the same set of works, then writes:

    site/data/publications.json  — list of {title, year, venue, type, doi}
    site/data/citations.json     — [{year, citations, cumulative}], plus total

Usage:
    python3 -X utf8 scripts/export_site_data.py [--orcid ORCID]
"""

import argparse
import json
from pathlib import Path

from openpyxl import load_workbook

from fetch_openalex import DEFAULT_ORCID, dedupe_preprints, fetch_works, is_excluded

ROOT = Path(__file__).resolve().parent.parent


def load_publications(xlsx_path: Path) -> list[dict]:
    wb = load_workbook(xlsx_path)
    ws = wb.active
    headers = [c.value for c in ws[1]]
    idx = {h: i for i, h in enumerate(headers)}

    pubs = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row[idx["Title"]]:
            continue
        pubs.append(
            {
                "title": row[idx["Title"]],
                "year": row[idx["Year"]],
                "venue": row[idx["Venue"]],
                "type": row[idx["Type"]],
                "doi": row[idx["DOI"]] if "DOI" in idx else "",
            }
        )
    pubs.sort(key=lambda p: (-(p["year"] or 0), p["title"]))
    return pubs


def citations_by_year(orcid: str) -> list[dict]:
    works = fetch_works(orcid)
    works = [w for w in works if not is_excluded(w.get("title", ""))]
    works = dedupe_preprints(works)

    totals: dict[int, int] = {}
    for w in works:
        for entry in w.get("counts_by_year", []):
            year = entry["year"]
            totals[year] = totals.get(year, 0) + entry["cited_by_count"]

    years = range(min(totals), max(totals) + 1) if totals else []
    series = []
    cumulative = 0
    for year in years:
        count = totals.get(year, 0)
        cumulative += count
        series.append({"year": year, "citations": count, "cumulative": cumulative})
    return series


def main() -> None:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("--orcid", default=DEFAULT_ORCID)
    args = parser.parse_args()

    pubs = load_publications(ROOT / "data" / "publications.xlsx")
    citations = citations_by_year(args.orcid)

    out_dir = ROOT / "site" / "data"
    out_dir.mkdir(parents=True, exist_ok=True)

    (out_dir / "publications.json").write_text(
        json.dumps(pubs, indent=2, ensure_ascii=False), encoding="utf-8"
    )
    (out_dir / "citations.json").write_text(
        json.dumps(citations, indent=2, ensure_ascii=False), encoding="utf-8"
    )

    print(f"Wrote {len(pubs)} publications and {len(citations)} citation-years to {out_dir}")


if __name__ == "__main__":
    main()
