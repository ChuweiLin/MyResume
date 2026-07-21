#!/usr/bin/env python3
"""Fetch OpenAlex data and refresh publications.xlsx + the site's JSON data.

Run this whenever you want the site to reflect the latest OpenAlex records:

    python3 -X utf8 scripts/sync_publications.py [--orcid ORCID] [--xlsx data/publications.xlsx]

Every run:
  1. Fetches all works for --orcid from OpenAlex (excluding known name
     collisions in EXCLUDE_TITLES, collapsing preprint/published duplicate
     pairs down to the published version).
  2. Creates data/publications.xlsx from scratch if it doesn't exist yet, or
     otherwise refreshes only its DOI column — any manual edits to title,
     venue, type, or year are left untouched.
  3. Writes site/data/publications.json and site/data/citations.json from
     the spreadsheet and OpenAlex's per-year citation counts.
"""

import argparse
import json
import re
import ssl
import urllib.request
from pathlib import Path

import certifi
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter

ROOT = Path(__file__).resolve().parent.parent
DEFAULT_ORCID = "0000-0002-8640-5695"
DEFAULT_XLSX = ROOT / "data" / "publications.xlsx"
FONT_NAME = "Arial"

# Title fragments (lowercased) confirmed to belong to a different person with
# the same name, not the ORCID holder.
EXCLUDE_TITLES = [
    "clinical efficacy of cryopreserved autologous bone flaps",
    "association between childhood trauma and depression",
]


def fetch_works(orcid: str) -> list[dict]:
    url = (
        "https://api.openalex.org/works"
        f"?filter=author.orcid:{orcid}&per-page=100&sort=publication_year:desc"
    )
    req = urllib.request.Request(url, headers={"User-Agent": "MyPublications-script/1.0"})
    ctx = ssl.create_default_context(cafile=certifi.where())
    with urllib.request.urlopen(req, context=ctx) as resp:
        data = json.load(resp)
    return data["results"]


def normalize_title(title: str) -> str:
    return re.sub(r"[^a-z0-9]+", " ", (title or "").lower()).strip()


def is_excluded(title: str) -> bool:
    norm = normalize_title(title)
    return any(frag in norm for frag in EXCLUDE_TITLES)


def dedupe_preprints(works: list[dict]) -> list[dict]:
    by_title: dict[str, list[dict]] = {}
    for w in works:
        by_title.setdefault(normalize_title(w.get("title")), []).append(w)

    result = []
    for group in by_title.values():
        if len(group) == 1:
            result.append(group[0])
            continue
        published = [w for w in group if w.get("type") != "preprint"]
        result.append(published[0] if published else group[0])
    return result


def venue_name(work: dict) -> str:
    loc = work.get("primary_location") or {}
    source = loc.get("source") or {}
    return source.get("display_name") or ""


def doi(work: dict) -> str:
    return work.get("doi") or work.get("ids", {}).get("doi") or ""


def coauthors(work: dict, orcid: str) -> str:
    orcid_url = f"https://orcid.org/{orcid}"
    names = [
        a.get("author", {}).get("display_name", "")
        for a in work.get("authorships", [])
        if a.get("author", {}).get("orcid") != orcid_url
    ]
    return ", ".join(n for n in names if n)


def build_rows(works: list[dict], orcid: str) -> list[tuple]:
    rows = []
    for w in works:
        rows.append(
            (
                w.get("title") or "",
                w.get("publication_year"),
                venue_name(w),
                w.get("type") or "",
                coauthors(w, orcid),
                doi(w),
            )
        )
    rows.sort(key=lambda r: (-(r[1] or 0), r[0]))
    return rows


def create_spreadsheet(works: list[dict], orcid: str, out_path: Path) -> None:
    rows = build_rows(works, orcid)

    wb = Workbook()
    ws = wb.active
    ws.title = "Publications"

    headers = ["Title", "Year", "Venue", "Type", "Co-authors", "DOI"]
    for col, header in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = Font(name=FONT_NAME, bold=True)

    for row_idx, row in enumerate(rows, start=2):
        for col_idx, value in enumerate(row, start=1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.font = Font(name=FONT_NAME)

    widths = [70, 8, 45, 16, 60, 38]
    for col, width in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(col)].width = width

    ws.freeze_panes = "A2"

    out_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(out_path)
    print(f"Created {out_path} with {len(rows)} publications")


def refresh_doi_column(works: list[dict], xlsx_path: Path) -> None:
    doi_map = {normalize_title(w.get("title")): doi(w) for w in works}

    wb = load_workbook(xlsx_path)
    ws = wb.active

    headers = [c.value for c in ws[1]]
    if "DOI" in headers:
        doi_col = headers.index("DOI") + 1
    else:
        doi_col = len(headers) + 1
        cell = ws.cell(row=1, column=doi_col, value="DOI")
        cell.font = Font(name=FONT_NAME, bold=True)
        ws.column_dimensions[cell.column_letter].width = 38

    title_col = headers.index("Title") + 1

    unmatched = []
    filled = 0
    for row in ws.iter_rows(min_row=2):
        title = row[title_col - 1].value
        match = doi_map.get(normalize_title(title))
        cell = ws.cell(row=row[0].row, column=doi_col)
        if match:
            cell.value = match
            cell.font = Font(name=FONT_NAME)
            filled += 1
        elif not cell.value:
            unmatched.append(title)

    wb.save(xlsx_path)
    print(f"Refreshed DOI for {filled} rows in {xlsx_path}")
    if unmatched:
        print("No OpenAlex match found for:")
        for t in unmatched:
            print(f"  - {t}")


def load_publications(xlsx_path: Path) -> list[dict]:
    wb = load_workbook(xlsx_path)
    ws = wb.active
    headers = [c.value for c in ws[1]]
    idx = {h: i for i, h in enumerate(headers)}

    pubs = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row[idx["Title"]]:
            continue
        pubs.append(
            {
                "title": row[idx["Title"]],
                "year": row[idx["Year"]],
                "venue": row[idx["Venue"]],
                "type": row[idx["Type"]],
                "doi": row[idx["DOI"]] if "DOI" in idx else "",
            }
        )
    pubs.sort(key=lambda p: (-(p["year"] or 0), p["title"]))
    return pubs


def citations_by_year(works: list[dict]) -> list[dict]:
    totals: dict[int, int] = {}
    for w in works:
        for entry in w.get("counts_by_year", []):
            year = entry["year"]
            totals[year] = totals.get(year, 0) + entry["cited_by_count"]

    years = range(min(totals), max(totals) + 1) if totals else []
    series = []
    cumulative = 0
    for year in years:
        count = totals.get(year, 0)
        cumulative += count
        series.append({"year": year, "citations": count, "cumulative": cumulative})
    return series


def export_site_data(pubs: list[dict], citations: list[dict]) -> None:
    out_dir = ROOT / "site" / "data"
    out_dir.mkdir(parents=True, exist_ok=True)

    (out_dir / "publications.json").write_text(
        json.dumps(pubs, indent=2, ensure_ascii=False), encoding="utf-8"
    )
    (out_dir / "citations.json").write_text(
        json.dumps(citations, indent=2, ensure_ascii=False), encoding="utf-8"
    )
    print(f"Wrote {len(pubs)} publications and {len(citations)} citation-years to {out_dir}")


def main() -> None:
    parser = argparse.ArgumentParser(
        description=__doc__, formatter_class=argparse.RawDescriptionHelpFormatter
    )
    parser.add_argument("--orcid", default=DEFAULT_ORCID)
    parser.add_argument("--xlsx", default=str(DEFAULT_XLSX))
    args = parser.parse_args()

    xlsx_path = Path(args.xlsx)

    works = fetch_works(args.orcid)
    works = [w for w in works if not is_excluded(w.get("title", ""))]
    works = dedupe_preprints(works)

    if xlsx_path.exists():
        refresh_doi_column(works, xlsx_path)
    else:
        create_spreadsheet(works, args.orcid, xlsx_path)

    pubs = load_publications(xlsx_path)
    citations = citations_by_year(works)
    export_site_data(pubs, citations)


if __name__ == "__main__":
    main()
