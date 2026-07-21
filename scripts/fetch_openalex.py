#!/usr/bin/env python3
"""Fetch a researcher's works from OpenAlex by ORCID and save them to a spreadsheet.

Usage:
    python3 -X utf8 scripts/fetch_openalex.py [--orcid ORCID] [--out data/publications.xlsx]

Excludes works whose title matches an entry in EXCLUDE_TITLES (papers that share
the ORCID holder's name but belong to a different person), and collapses
preprint/published duplicate pairs down to the published version.
"""

import argparse
import json
import re
import ssl
import urllib.request
from pathlib import Path

import certifi
from openpyxl import Workbook
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter

DEFAULT_ORCID = "0000-0002-8640-5695"

# Title fragments (lowercased) confirmed to belong to a different person with
# the same name, not the ORCID holder.
EXCLUDE_TITLES = [
    "clinical efficacy of cryopreserved autologous bone flaps",
    "association between childhood trauma and depression",
]

FONT_NAME = "Arial"


def fetch_works(orcid: str) -> list[dict]:
    url = (
        "https://api.openalex.org/works"
        f"?filter=author.orcid:{orcid}&per-page=100&sort=publication_year:desc"
    )
    req = urllib.request.Request(url, headers={"User-Agent": "MyPublications-script/1.0"})
    ctx = ssl.create_default_context(cafile=certifi.where())
    with urllib.request.urlopen(req, context=ctx) as resp:
        data = json.load(resp)
    return data["results"]


def normalize_title(title: str) -> str:
    return re.sub(r"[^a-z0-9]+", " ", (title or "").lower()).strip()


def is_excluded(title: str) -> bool:
    norm = normalize_title(title)
    return any(frag in norm for frag in EXCLUDE_TITLES)


def dedupe_preprints(works: list[dict]) -> list[dict]:
    by_title: dict[str, list[dict]] = {}
    for w in works:
        by_title.setdefault(normalize_title(w.get("title")), []).append(w)

    result = []
    for group in by_title.values():
        if len(group) == 1:
            result.append(group[0])
            continue
        published = [w for w in group if w.get("type") != "preprint"]
        result.append(published[0] if published else group[0])
    return result


def venue_name(work: dict) -> str:
    loc = work.get("primary_location") or {}
    source = loc.get("source") or {}
    return source.get("display_name") or ""


def coauthors(work: dict, orcid: str) -> str:
    orcid_url = f"https://orcid.org/{orcid}"
    names = [
        a.get("author", {}).get("display_name", "")
        for a in work.get("authorships", [])
        if a.get("author", {}).get("orcid") != orcid_url
    ]
    return ", ".join(n for n in names if n)


def build_rows(works: list[dict], orcid: str) -> list[tuple]:
    rows = []
    for w in works:
        rows.append(
            (
                w.get("title") or "",
                w.get("publication_year"),
                venue_name(w),
                w.get("type") or "",
                coauthors(w, orcid),
            )
        )
    rows.sort(key=lambda r: (-(r[1] or 0), r[0]))
    return rows


def write_spreadsheet(rows: list[tuple], out_path: Path) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "Publications"

    headers = ["Title", "Year", "Venue", "Type", "Co-authors"]
    for col, header in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = Font(name=FONT_NAME, bold=True)

    for row_idx, row in enumerate(rows, start=2):
        for col_idx, value in enumerate(row, start=1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.font = Font(name=FONT_NAME)

    widths = [70, 8, 45, 16, 60]
    for col, width in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(col)].width = width

    ws.freeze_panes = "A2"

    out_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(out_path)


def main() -> None:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("--orcid", default=DEFAULT_ORCID)
    parser.add_argument(
        "--out",
        default=str(Path(__file__).resolve().parent.parent / "data" / "publications.xlsx"),
    )
    args = parser.parse_args()

    works = fetch_works(args.orcid)
    works = [w for w in works if not is_excluded(w.get("title", ""))]
    works = dedupe_preprints(works)
    rows = build_rows(works, args.orcid)

    out_path = Path(args.out)
    write_spreadsheet(rows, out_path)

    print(f"Saved {len(rows)} publications to {out_path}")


if __name__ == "__main__":
    main()
