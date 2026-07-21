#!/usr/bin/env python3
"""Add or refresh a DOI column on an existing publications spreadsheet.

Matches each row to an OpenAlex work by (normalized) title and fills in a
"DOI" column, without touching any other column or row — safe to run after
you've hand-edited the spreadsheet. Rows with no OpenAlex match are left
blank and reported on stdout.

Usage:
    python3 -X utf8 scripts/add_doi_column.py [--orcid ORCID] [--file data/publications.xlsx]
"""

import argparse
from pathlib import Path

from openpyxl import load_workbook
from openpyxl.styles import Font

from fetch_openalex import (
    DEFAULT_ORCID,
    FONT_NAME,
    dedupe_preprints,
    doi,
    fetch_works,
    is_excluded,
    normalize_title,
)


def build_doi_map(orcid: str) -> dict[str, str]:
    works = fetch_works(orcid)
    works = [w for w in works if not is_excluded(w.get("title", ""))]
    works = dedupe_preprints(works)
    return {normalize_title(w.get("title")): doi(w) for w in works}


def main() -> None:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("--orcid", default=DEFAULT_ORCID)
    parser.add_argument(
        "--file",
        default=str(Path(__file__).resolve().parent.parent / "data" / "publications.xlsx"),
    )
    args = parser.parse_args()

    doi_map = build_doi_map(args.orcid)

    path = Path(args.file)
    wb = load_workbook(path)
    ws = wb.active

    headers = [c.value for c in ws[1]]
    if "DOI" in headers:
        doi_col = headers.index("DOI") + 1
    else:
        doi_col = len(headers) + 1
        cell = ws.cell(row=1, column=doi_col, value="DOI")
        cell.font = Font(name=FONT_NAME, bold=True)
        ws.column_dimensions[cell.column_letter].width = 38

    title_col = headers.index("Title") + 1

    unmatched = []
    filled = 0
    for row in ws.iter_rows(min_row=2):
        title = row[title_col - 1].value
        match = doi_map.get(normalize_title(title))
        cell = ws.cell(row=row[0].row, column=doi_col)
        if match:
            cell.value = match
            cell.font = Font(name=FONT_NAME)
            filled += 1
        elif not cell.value:
            unmatched.append(title)

    wb.save(path)
    print(f"Filled DOI for {filled} rows in {path}")
    if unmatched:
        print("No OpenAlex match found for:")
        for t in unmatched:
            print(f"  - {t}")


if __name__ == "__main__":
    main()
